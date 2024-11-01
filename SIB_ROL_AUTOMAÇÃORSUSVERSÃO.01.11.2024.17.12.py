import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import platform
import customtkinter as ctk
from tkinter import filedialog, messagebox

# Função para obter a lista de arquivos .txt da pasta SIB
def obter_arquivos_sib(diretorio_sib):
    arquivos_txt = []
    for arquivo in os.listdir(diretorio_sib):
        if arquivo.endswith('.txt'):
            arquivos_txt.append(os.path.join(diretorio_sib, arquivo))
    return arquivos_txt

# Função para mapear o mês e ano do atendimento ao arquivo SIB correto
def obter_arquivo_sib(data_inicio_atendimento, arquivos_txt):
    mes_ano_inicio = data_inicio_atendimento.strftime('%m%Y')
    for arquivo_txt in arquivos_txt:
        if mes_ano_inicio in arquivo_txt:
            return arquivo_txt
    return None

# Função para verificar a cobertura do beneficiário
def verificar_cobertura(data_contratacao, data_cancelamento, data_reativacao, data_inicio_atendimento):
    # Verifica se o beneficiário foi reativado após o cancelamento
    if pd.notna(data_cancelamento) and pd.notna(data_reativacao) and data_cancelamento > data_reativacao:
        if data_inicio_atendimento > data_cancelamento:
            return 'INATIVO', "Início do atendimento após data de cancelamento.", ""

    # Caso em que não há cancelamento nem reativação
    if pd.isna(data_cancelamento) and pd.isna(data_reativacao):
        if pd.notnull(data_contratacao) and data_contratacao <= data_inicio_atendimento:
            # Verificação de carência
            if data_inicio_atendimento <= (data_contratacao + pd.DateOffset(months=6)):
                return 'ATIVO', "Sem cancelamento/reativação, dentro do período de carência.", "Possível carência ou CPT"
            elif data_inicio_atendimento <= (data_contratacao + pd.DateOffset(years=2)):
                return 'ATIVO', "Sem cancelamento/reativação, dentro do período de CPT.", "Possível CPT"
            else:
                return 'ATIVO', "Sem cancelamento/reativação, fora dos períodos de carência ou CPT.", ""

    # Caso em que há reativação
    elif pd.notna(data_reativacao):
        if data_reativacao <= data_inicio_atendimento:
            # Reinicia a contagem de carência
            if pd.notna(data_cancelamento) and data_cancelamento < data_reativacao - pd.DateOffset(days=30):
                # Verificação de carência a partir da data de reativação
                if data_inicio_atendimento <= (data_reativacao + pd.DateOffset(months=6)):
                    return 'ATIVO', "Reativado, dentro do período de carência.", "Possível carência ou CPT"
                elif data_inicio_atendimento <= (data_reativacao + pd.DateOffset(years=2)):
                    return 'ATIVO', "Reativado, dentro do período de CPT.", "Possível CPT"
                else:
                    return 'ATIVO', "Reativado, fora dos períodos de carência ou CPT.", ""
            else:
                # Aqui verificamos a carência e CPT a partir da data de contratação
                if data_inicio_atendimento <= (data_contratacao + pd.DateOffset(months=6)):
                    return 'ATIVO', "Reativado em menos de 30 dias após o cancelamento, dentro do período de carência.", "Possível carência ou CPT"
                elif data_inicio_atendimento <= (data_contratacao + pd.DateOffset(years=2)):
                    return 'ATIVO', "Reativado em menos de 30 dias após o cancelamento, dentro do período de CPT.", "Possível CPT"
                else:
                    return 'ATIVO', "Reativado em menos de 30 dias após o cancelamento, fora dos períodos de carência ou CPT.", ""

    # Caso geral para verificação de contrato
    elif pd.notnull(data_contratacao) and data_contratacao <= data_inicio_atendimento and (pd.isna(data_cancelamento) or data_inicio_atendimento <= data_cancelamento):
        # Verificação de carência
        if data_inicio_atendimento <= (data_contratacao + pd.DateOffset(months=6)):
            return 'ATIVO', "Contrato válido e dentro do período de carência.", "Possível carência ou CPT"
        elif data_inicio_atendimento <= (data_contratacao + pd.DateOffset(years=2)):
            return 'ATIVO', "Contrato válido e dentro do período de CPT.", "Possível CPT"
        else:
            return 'ATIVO', "Contrato válido e fora dos períodos de carência ou CPT.", ""

    return 'INATIVO', "Nenhuma condição satisfeita.", ""


def processar_arquivo_abi(df_abi, arquivos_txt):
    df_abi['situacao'] = 'INATIVO'
    df_abi['situacao_carencia_cpt'] = ''
    beneficiarios_processados = set()

    for index, row in df_abi.iterrows():
        data_inicio_atendimento = pd.to_datetime(row['dataInicioAtendimento'], errors='coerce')
        codigo_beneficiario = str(row['codigoBeneficiario'])

        arquivo_sib = obter_arquivo_sib(data_inicio_atendimento, arquivos_txt)

        if arquivo_sib is not None:
            df_sib_txt = pd.read_csv(arquivo_sib, sep='\t')  # Ajuste o delimitador conforme necessário
            beneficiario_sib = df_sib_txt[df_sib_txt['codigoBeneficiario'].astype(str) == codigo_beneficiario]

            if not beneficiario_sib.empty:
                situacao = 'INATIVO'
                mensagem = ""
                situacao_carencia_cpt = ""
                for _, sib_row in beneficiario_sib.iterrows():
                    data_contratacao = pd.to_datetime(sib_row['dataContratacao'], errors='coerce')
                    data_cancelamento = pd.to_datetime(sib_row['dataCancelamento'], errors='coerce')
                    data_reativacao = pd.to_datetime(sib_row['dataReativacao'], errors='coerce')

                    situacao, mensagem, situacao_carencia_cpt = verificar_cobertura(
                        data_contratacao, data_cancelamento, data_reativacao, data_inicio_atendimento)

                    if codigo_beneficiario not in beneficiarios_processados and data_inicio_atendimento not in beneficiarios_processados:
                        print(f"\n=== Beneficiário: {codigo_beneficiario} ===")
                        print(f"  - Data de Contratação: {data_contratacao}")
                        print(f"  - Data de Cancelamento: {data_cancelamento}")
                        print(f"  - Data de Reativação: {data_reativacao}")
                        print(f"  - Início do Atendimento: {data_inicio_atendimento}")
                        print(f"  - Situação: {situacao} - {mensagem}")
                        print(f"  - Situação de Carência/CPT: {situacao_carencia_cpt}")
                        beneficiarios_processados.add(codigo_beneficiario)
                        beneficiarios_processados.add(data_inicio_atendimento)

                df_abi.at[index, 'situacao'] = situacao
                df_abi.at[index, 'situacao_carencia_cpt'] = situacao_carencia_cpt
            else:
                print(f"Nenhum registro SIB encontrado para beneficiário {codigo_beneficiario}.")
        else:
            print(f"Arquivo SIB não encontrado para a data {data_inicio_atendimento.strftime('%m%Y')}")

    return df_abi


# Funções do segundo programa
def carregar_planilhas(mapeamento_path, correlacao_path):
    mapeamento_df = pd.read_excel(mapeamento_path, sheet_name='Mapeamento ativos')
    correlacao_df = pd.read_excel(correlacao_path)
    return mapeamento_df, correlacao_df

def filtrar_mapeamento(mapeamento_df):
    return mapeamento_df[~mapeamento_df['Grau de equivalencia '].isin(['', ''])]

def criar_dicionarios(mapeamento_df, correlacao_df):
    # Criar um dicionário que mapeia cada código Sigtap a uma lista de tuplas (Código TUSS, Grau de Equivalência)
    sigTap_to_tuss = {}
    for _, row in mapeamento_df.iterrows():
        sigTap_code = row['Código Sigtap Final']
        sigTap_to_tuss.setdefault(sigTap_code, []).append((row['Código TUSS'], row['Grau de equivalencia ']))

    tuss_cobertura = correlacao_df.set_index(correlacao_df.columns[0])[correlacao_df.columns[2]].to_dict()
    
    return sigTap_to_tuss, tuss_cobertura
def mapear_codigo_tuss(rsus_df, sigTap_to_tuss):
    def mapear(codigo_proced):
        if pd.isna(codigo_proced) or codigo_proced == '':
            return 'código proced vazio'
        elif codigo_proced in sigTap_to_tuss:
            return 'mapeado'
        else:
            return 'não mapeado'

    rsus_df['statusMapeamento'] = rsus_df['codigoProcedimento'].apply(mapear)
    
    # Adiciona o código TUSS e o Grau de Equivalência
    rsus_df['codigoTUSS'] = rsus_df.apply(
        lambda row: sigTap_to_tuss.get(row['codigoProcedimento'], [('não encontrado', 'não encontrado')])[0][0]
                     if row['statusMapeamento'] == 'mapeado' else 'não encontrado',
        axis=1
    )
    
    # Adiciona a coluna de Grau de Equivalência
    rsus_df['Grau de Equivalencia'] = rsus_df.apply(
        lambda row: sigTap_to_tuss.get(row['codigoProcedimento'], [('não encontrado', 'não encontrado')])[0][1]
                     if row['statusMapeamento'] == 'mapeado' else 'não encontrado',
        axis=1
    )



def verificar_coberturaTUSS(rsus_df, tuss_cobertura):
    rsus_df['CoberturaObrigatoria'] = rsus_df['codigoTUSS'].apply(
        lambda x: "SIM" if x in tuss_cobertura and tuss_cobertura[x] == "SIM" else "NÃO"
    )

def gerar_relatorio(rsus_df):
    for codigo in rsus_df['codigoBeneficiario'].unique():
        print(f"Código Beneficiário: {codigo}\n")
        dados_beneficiario = rsus_df[rsus_df['codigoBeneficiario'] == codigo]
        for index, row in dados_beneficiario.iterrows():
            print(
                f"  Código Procedimento: {row['codigoProcedimento']}\n"
                f"  Status Mapeamento: {row['statusMapeamento']}\n"
                f"  Código TUSS: {row['codigoTUSS']}\n"
                f"  Cobertura Obrigatória: {row['CoberturaObrigatoria']}\n"
            )
        print("\n")  # Espaço extra entre beneficiários

def salvar_excel(rsus_df, output_path):
    rsus_df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active
    preenchimento = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for row in range(2, len(rsus_df) + 2):
        if ws[f'E{row}'].value == "NÃO":
            for cell in ws[row]:
                cell.fill = preenchimento

    if platform.system() == "Windows":
        os.startfile(output_path)
    elif platform.system() == "Darwin":  # MacOS
        os.system(f"open {output_path}")
    else:  # Assume Linux or other Unix-like system
        os.system(f"xdg-open {output_path}")

    wb.save(output_path)

def gerar_resumo_beneficiarios(rsus_df):
    resumo = []
    for codigo in rsus_df['codigoBeneficiario'].unique():
        dados_beneficiario = rsus_df[rsus_df['codigoBeneficiario'] == codigo]
        situacao_geral = dados_beneficiario['situacao'].unique()
        if len(situacao_geral) > 1:
            situacao_geral = 'MISTO (verifique individualmente)'
        else:
            situacao_geral = situacao_geral[0]

        procedimentos = len(dados_beneficiario)

        tuss_cobertura = [
            f"{str(tuss)} ({cobertura})" 
            for tuss, cobertura in zip(dados_beneficiario['codigoTUSS'].astype(str), dados_beneficiario['CoberturaObrigatoria'])
        ]
        
        resumo.append({
            'Código Beneficiário': codigo,
            'Total de Procedimentos': procedimentos,
            'Situação Geral': situacao_geral,
            'Códigos TUSS': ', '.join(tuss_cobertura),
        })

    resumo_df = pd.DataFrame(resumo)
    resumo_df.to_excel('resumo_beneficiarios.xlsx', index=False)
    return resumo_df

# Função para gerar a interface gráfica com customtkinter
def criar_interface():
    app = ctk.CTk()
    app.geometry("500x450")
    app.title("Sistema de Análise de Cobertura e Procedimentos")

    def escolher_arquivo_abi():
        caminho_arquivo_abi = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
        entry_arquivo_abi.insert(0, caminho_arquivo_abi)

    def escolher_pasta_sib():
        caminho_pasta_sib = filedialog.askdirectory()
        entry_pasta_sib.insert(0, caminho_pasta_sib)

    def escolher_mapeamento():
        caminho_mapeamento = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
        entry_mapeamento.insert(0, caminho_mapeamento)

    def escolher_correlacao():
        caminho_correlacao = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
        entry_correlacao.insert(0, caminho_correlacao)

    def escolher_saida():
        caminho_saida = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
        entry_saida.insert(0, caminho_saida)

    def processar():
        caminho_arquivo_abi = entry_arquivo_abi.get()
        caminho_pasta_sib = entry_pasta_sib.get()
        caminho_mapeamento = entry_mapeamento.get()
        caminho_correlacao = entry_correlacao.get()
        caminho_saida = entry_saida.get()

        if not all([caminho_arquivo_abi, caminho_pasta_sib, caminho_mapeamento, caminho_correlacao, caminho_saida]):
            messagebox.showwarning("Campos Faltando", "Preencha todos os campos.")
            return

        df_abi = pd.read_excel(caminho_arquivo_abi)
        arquivos_txt = obter_arquivos_sib(caminho_pasta_sib)
        df_abi = processar_arquivo_abi(df_abi, arquivos_txt)

        mapeamento_df, correlacao_df = carregar_planilhas(caminho_mapeamento, caminho_correlacao)
        mapeamento_df = filtrar_mapeamento(mapeamento_df)
        sigTap_to_tuss, tuss_cobertura = criar_dicionarios(mapeamento_df, correlacao_df)

        mapear_codigo_tuss(df_abi, sigTap_to_tuss)
        verificar_coberturaTUSS(df_abi, tuss_cobertura)

        salvar_excel(df_abi, caminho_saida)

        gerar_relatorio(df_abi)
        resumo_df = gerar_resumo_beneficiarios(df_abi)

        messagebox.showinfo("Processamento Completo", f"Processamento finalizado. Resumo salvo em {caminho_saida}")

    frame = ctk.CTkFrame(app)
    frame.pack(pady=20, padx=20, fill="both", expand=True)

    label_arquivo_abi = ctk.CTkLabel(frame, text="Arquivo ABI:")
    label_arquivo_abi.grid(row=0, column=0, padx=10, pady=10, sticky="w")
    entry_arquivo_abi = ctk.CTkEntry(frame)
    entry_arquivo_abi.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
    btn_arquivo_abi = ctk.CTkButton(frame, text="Selecionar", command=escolher_arquivo_abi)
    btn_arquivo_abi.grid(row=0, column=2, padx=10, pady=10)

    label_pasta_sib = ctk.CTkLabel(frame, text="Pasta SIB:")
    label_pasta_sib.grid(row=1, column=0, padx=10, pady=10, sticky="w")
    entry_pasta_sib = ctk.CTkEntry(frame)
    entry_pasta_sib.grid(row=1, column=1, padx=10, pady=10, sticky="ew")
    btn_pasta_sib = ctk.CTkButton(frame, text="Selecionar", command=escolher_pasta_sib)
    btn_pasta_sib.grid(row=1, column=2, padx=10, pady=10)

    label_mapeamento = ctk.CTkLabel(frame, text="Mapeamento TUSS:")
    label_mapeamento.grid(row=2, column=0, padx=10, pady=10, sticky="w")
    entry_mapeamento = ctk.CTkEntry(frame)
    entry_mapeamento.grid(row=2, column=1, padx=10, pady=10, sticky="ew")
    btn_mapeamento = ctk.CTkButton(frame, text="Selecionar", command=escolher_mapeamento)
    btn_mapeamento.grid(row=2, column=2, padx=10, pady=10)

    label_correlacao = ctk.CTkLabel(frame, text="Correlacao TUSS Rol:")
    label_correlacao.grid(row=3, column=0, padx=10, pady=10, sticky="w")
    entry_correlacao = ctk.CTkEntry(frame)
    entry_correlacao.grid(row=3, column=1, padx=10, pady=10, sticky="ew")
    btn_correlacao = ctk.CTkButton(frame, text="Selecionar", command=escolher_correlacao)
    btn_correlacao.grid(row=3, column=2, padx=10, pady=10)

    label_saida = ctk.CTkLabel(frame, text="Local de Saída:")
    label_saida.grid(row=4, column=0, padx=10, pady=10, sticky="w")
    entry_saida = ctk.CTkEntry(frame)
    entry_saida.grid(row=4, column=1, padx=10, pady=10, sticky="ew")
    btn_saida = ctk.CTkButton(frame, text="Selecionar", command=escolher_saida)
    btn_saida.grid(row=4, column=2, padx=10, pady=10)

    btn_processar = ctk.CTkButton(frame, text="Processar", command=processar)
    btn_processar.grid(row=5, column=0, columnspan=3, padx=10, pady=20, sticky="ew")

    app.mainloop()

criar_interface()




